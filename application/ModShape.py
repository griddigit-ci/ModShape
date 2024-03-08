# Licensed under the Mozilla Public License 2.0.
# Copyright(c) 2023, gridDigIt Kft. All rights reserved.
# author: Chavdar Ivanov

import io
import os
import sys
import time
import zipfile
import requests
import rdflib.term
from PyQt5.QtWidgets import *
from pyshacl import validate
from rdflib import Graph, RDFS, RDF, OWL
import pyoxigraph
from application import gui
import multiprocessing
from itertools import islice, tee
import polars as pl
from numba import njit
from joblib import Parallel, delayed
from concurrent.futures import ThreadPoolExecutor
from itertools import zip_longest
from toolz import pipe, map
import lightrdf
from itertools import islice
import re
import concurrent.futures
import openpyxl
from openpyxl import Workbook
from application.CimGraph import CimGraph
from collections import OrderedDict
from openpyxl.styles import Alignment


class ModShape(QDialog, gui.Ui_Dialog):
    datatypes_mapping = pl.DataFrame()

    def __init__(self, parent=None):
        super(ModShape, self).__init__(parent)
        self.setupUi(self)
        global project_path

        project_path = ""

        self.file_types = {
            'Instance Data': {'filter': "Instance data, (*.xml *.zip *.cimx)", 'button': self.InstanceData,
                              'label': self.labelSelectData},
            'SHACL Constraints': {'filter': "SHACL Constraints, *.ttl", 'button': self.SHACLConstraints,
                                  'label': self.labelSelectDataSHACL},
            'RDF Datatypes': {'filter': "RDFS datatypes, *.xlsx", 'button': self.RDFDatatypes,
                              'label': self.labelSelectRDFmap}
        }

        for action, details in self.file_types.items():
            self.connect_button(details['button'], action, details['filter'], details['label'])

        self.buttonOK.clicked.connect(lambda: self.push_button_ok())

        self.merged_instance_graph = Graph()
        self.merged_shacl_graph = Graph()
        # self.merged_datatype_graph = Graph()
        # self.datatypes_mapping = pl.DataFrame()

    def connect_button(self, button, action, file_filter, label):
        button.clicked.connect(lambda: self.select_file(action, project_path, file_filter, label))

    def select_file(self, action, pathi, file_filter, letext):
        selection = QFileDialog.getOpenFileNames(self, f"Please select the {action} file", pathi, file_filter)[0]

        if selection:
            letext.setText('\n'.join(selection))
            # Store the selection in a dictionary or any other data structure
            self.file_types[action]['selection'] = selection

    @staticmethod
    def get_format_from_extension(extension):
        # Map file extensions to RDF serialization formats
        extension_format_mapping = {
            '.xml': 'xml',
            '.rdf': 'xml',
            '.ttl': 'turtle',
            # Add more mappings as needed
        }
        return extension_format_mapping.get(extension, None)

    @staticmethod
    def is_supported_archive(extension):
        # Check if the file extension corresponds to a supported archive format
        supported_archive_extensions = ['.zip', '.cimx']  # Add more extensions as needed
        return extension.lower() in supported_archive_extensions

    def process_iterator(self, pyoxigraph_iterator):
        # local_graph_data = []
        local_graph_data = Graph()

        for s, p, o in pyoxigraph_iterator:
            graph_subject = rdflib.URIRef(s.value)
            graph_predicate = rdflib.URIRef(p.value)

            if hasattr(o, 'datatype'):  # do literal
                # the way to get the datatype
                # local_datatype_mapping_graph.row(None,by_predicate=pl.col("Property").str.contains("http://iec.ch/TC57/2013/CIM-schema-cim16#UnderexcitationLimiterUserDefined.proprietary"))[1]
                # datatype_from_map = self.datatypes_mapping.row(None, by_predicate=pl.col("Property").str.find(p.value))[1]
                datatype_from_map = self.datatypes_mapping.filter(pl.col("Property") == p.value)
                if datatype_from_map.is_empty():
                    graph_object = rdflib.Literal(o.value, datatype=rdflib.URIRef(o.datatype.value), lang=o.language)
                else:
                    graph_object = rdflib.Literal(o.value, datatype=rdflib.URIRef(datatype_from_map[0, 1].__str__()),
                                                  lang=o.language)
            else:
                graph_object = rdflib.URIRef(o.value)

            # local_graph_data.append((rdflib.IdentifiedNode(graph_subject), rdflib.IdentifiedNode(graph_predicate),
            #                          rdflib.IdentifiedNode(graph_object), rdflib.IdentifiedNode('http://hello.eu/modshape/test')))

            local_graph_data.add((graph_subject, graph_predicate, graph_object))

        return local_graph_data

    def process_shacl_iterator(self, pyoxigraph_shacl_iterator, local_shacl_graph_data):
        # local_shacl_graph_data = []

        for s, p, o in pyoxigraph_shacl_iterator:
            if p.value == OWL.imports.__str__() and o.value != 'http://www.w3.org/ns/shacl#':
                file_path = o.value
                with open(file_path, 'rb') as file:
                    content = file.read()
                    sub_iterator = pyoxigraph.parse(input=content, mime_type="text/turtle")
                    local_shacl_graph_data = self.process_shacl_iterator(sub_iterator, local_shacl_graph_data)
                    local_shacl_graph_data = local_shacl_graph_data + local_shacl_graph_data
            else:
                graph_subject = rdflib.URIRef(s.value)
                graph_predicate = rdflib.URIRef(p.value)
                if hasattr(o, 'datatype'):  # do literal
                    # graph_object = rdflib.Literal(o.value, datatype=rdflib.URIRef(o.datatype.value), lang=o.language)
                    graph_object = rdflib.Literal(o.value, datatype=rdflib.URIRef(o.datatype.value))
                else:
                    graph_object = rdflib.URIRef(o.value)

                # local_shacl_graph_data.append((graph_subject, graph_predicate, rdflib.IdentifiedNode(graph_object),
                #                                rdflib.IdentifiedNode("http://test.eu/graph1")))

                local_shacl_graph_data.add((graph_subject, graph_predicate, graph_object))

        return local_shacl_graph_data

    def process_entry_content(self, entry_name, content):
        # Process the content and add triples to the merged graph
        # Determine the format based on the file extension
        _, ext = os.path.splitext(entry_name)
        ext = ext.lower()  # Convert to lowercase for case-insensitive comparison
        if self.is_supported_archive(ext):
            # If it's a supported archive format, treat it as a zip file
            with zipfile.ZipFile(io.BytesIO(content), 'r') as zip_ref:
                for member in zip_ref.infolist():
                    with zip_ref.open(member.filename) as inner_entry_file:
                        inner_content = inner_entry_file.read()
                        self.process_entry_content(member.filename, inner_content)
        else:
            # If it's not a supported archive format, treat it as a regular file
            file_format = self.get_format_from_extension(ext)
            if file_format:
                # local_graph = Graph()
                # start variant B
                # local_graph.parse(data=content, format=file_format) # this way of parsing is slow
                # end variant B

                # this is if we want a list of the triples. It is taking 1-3 more seconds
                # triples = list(pyoxigraph.parse(input=content, mime_type="application/rdf+xml", base_iri="http://iec.ch/TC57/2013/CIM-schema-cim16#"))
                start_time_parsing = time.time()  # start time of parsing

                # start variant A with parsing, serialisation, and parsing in rdflib. Still need to see if the local_graph is in a good shape
                # pygraph = pyoxigraph.parse(input=content, mime_type="application/rdf+xml",
                #                            base_iri="http://iec.ch/TC57/2013/CIM-schema-cim16#")  # this is quick around 5 sec
                # binary_stream: IO[bytes] = io.BytesIO() # this is quick
                # pyoxigraph.serialize(input=pygraph, output=binary_stream, mime_type="application/n-triples") # this is quick, but 1-3 sec more than the parsing
                # binary_stream.seek(0) # this is quick
                # subjects, predicates, objects = zip_longest(*pygraph)
                #
                # df  = (pl.DataFrame({'data': objects}))
                # object_iterator = iter(objects)
                # v, dt, lan = zip_longest(*object_iterator, fillvalue=None)
                # data_frame = pl.DataFrame(data=[subjects, predicates, objects], schema=['s', 'p', 'o'])
                # object_subjects, object_predicates, object_values = zip_longest(*objects)

                # data_frame = pl.DataFrame(binary_stream, schema=['triple'])
                # local_graph.parse(source=binary_stream,format="nt") #this line takes around 6 min
                # end variant A

                # create the store (if .load is used then it is slow - 4 min; if .bulk_load is used then it is very fast 4 sec)
                # graph_store = pyoxigraph.Store()
                # graph_store.bulk_load(input=content, mime_type="application/rdf+xml",
                #                       base_iri="http://iec.ch/TC57/2013/CIM-schema-cim16#",
                #                       to_graph=pyoxigraph.NamedNode("http://example.eu/g"))

                # new experiments

                # start_time_newparsing_l = time.time()  # end time parsing
                # # Modified Parser
                # #parser = lightrdf.Parser()
                # start_time_newparsing_lb = time.time()  # end time parsing
                # # local_graph_data = Graph()
                # # doc = lightrdf.RDFDocument(io.BytesIO(content), parser=lightrdf.xml.PatternParser,base_iri="http://iec.ch/TC57/CIM100")
                # # triples_iter = doc.search_triples(None, None, None)
                # # end_time_newparsing_lb = time.time()  # end time parsing
                # # elapsed_time_newparsing_lb = end_time_newparsing_lb - start_time_newparsing_lb
                # # print(f"before for loop: {elapsed_time_newparsing_lb} seconds")
                # # batch_size = 1000
                #
                # angle_bracket_pattern = re.compile(r'^<(.+)>$')
                #
                # local_graph_data = Graph()
                # doc = lightrdf.RDFDocument(io.BytesIO(content), parser=lightrdf.xml.PatternParser,
                #                            base_iri="http://iec.ch/TC57/CIM100")
                # triples_iter = doc.search_triples(None, None, None)
                # end_time_newparsing_lb = time.time()  # end time parsing
                # elapsed_time_newparsing_lb = end_time_newparsing_lb - start_time_newparsing_lb
                # print(f"before for loop: {elapsed_time_newparsing_lb} seconds")
                # batch_size = 10000
                #
                # angle_bracket_pattern = re.compile(r'^<(.+)>$')
                #
                # with concurrent.futures.ThreadPoolExecutor() as executor:
                #     futures = []
                #
                #     for batch in iter(lambda: list(islice(triples_iter, batch_size)), []):
                #         batch_futures = [executor.submit(self.process_triple, triple, angle_bracket_pattern) for triple
                #                          in batch]
                #         futures.extend(batch_futures)
                #
                #     for future in concurrent.futures.as_completed(futures):
                #         result = future.result()
                #         #local_graph_data.add(result)
                #

                # ens new experiments

                # for batch in iter(lambda: list(islice(triples_iter, batch_size)), []):
                #     for triple in batch:
                #         s, p, o = triple
                #         match = angle_bracket_pattern.match(o)
                #         match_p = angle_bracket_pattern.match(p).group(1)
                #         if match:
                #             graph_object = rdflib.URIRef(match.group(1))
                #         else:
                #             #o_mod = o[1:-1]
                #             filtered_df = self.datatypes_mapping.filter(self.datatypes_mapping['Property'] == match_p)
                #             #datatype_from_map = self.datatypes_mapping.filter(pl.col("Property") == p)
                #             if not filtered_df.is_empty():
                #                 datatype_from_map = filtered_df[0]['Datatype']
                #             else:
                #                 datatype_from_map = None
                #
                #             if datatype_from_map is not None:
                #                 graph_object = rdflib.Literal(o, datatype=rdflib.XSD.string)
                #             else:
                #                 graph_object = rdflib.Literal(o,
                #                                               datatype=rdflib.URIRef(str(datatype_from_map[0, 1])))
                #
                #         local_graph_data.add((rdflib.URIRef(angle_bracket_pattern.match(s).group(1)), rdflib.URIRef(match_p), graph_object))

                # for triple in doc.search_triples(None,None,None):
                #     #start_time_newparsing_lbloop = time.time()  # end time parsing
                #     s = triple[0]
                #     p = triple[1]
                #     o = triple[2]
                #     if o.startswith('<'):
                #         graph_object = rdflib.URIRef(o[1:-1])
                #     else:
                #         o_mod = o[1:-1]
                #         datatype_from_map = self.datatypes_mapping.filter(pl.col("Property") == p)
                #         if datatype_from_map.is_empty():
                #             graph_object = rdflib.Literal(o_mod, datatype=rdflib.XSD.string)
                #         else:
                #             graph_object = rdflib.Literal(o_mod,datatype=rdflib.URIRef(datatype_from_map[0, 1].__str__()))

                #     end_time_newparsing_lbloop = time.time()  # end time parsing
                #  local_graph_data.add((rdflib.URIRef(s[1:-1]), rdflib.URIRef(p[1:-1]), graph_object))
                #     end_time_newparsing_lbloopbg = time.time()  # end time parsing
                #     elapsed_time_newparsing_lbloopbg = end_time_newparsing_lbloop - start_time_newparsing_lbloop
                #     print(f"for loop iter: {elapsed_time_newparsing_lbloopbg} seconds")
                #     elapsed_time_newparsing_lbG = end_time_newparsing_lbloopbg - start_time_newparsing_lbloop
                #     print(f"loop with adding to graph: {elapsed_time_newparsing_lbG} seconds")

                # end_time_newparsing_l = time.time()  # end time parsing
                # elapsed_time_newparsing_l = end_time_newparsing_l - start_time_newparsing_l
                # print(f"New parsing time Light: {elapsed_time_newparsing_l} seconds")
                # sys.exit()

                # start_time_newparsing = time.time()  # start time parsing
                # Modified Parser RDFLIB
                instance_data_graph_local = CimGraph()
                instance_data_graph_local.parse(data=content, format="cimxml", datatype_mapping=self.datatypes_mapping)
                self.merged_instance_graph = self.merged_instance_graph + instance_data_graph_local
                # end_time_newparsing = time.time()  # end time parsing
                # elapsed_time_newparsing = end_time_newparsing - start_time_newparsing
                # print(f"New parsing time: {elapsed_time_newparsing} seconds")

                # # start Variant C
                # pyoxigraph_iterator = pyoxigraph.parse(input=content, mime_type="application/rdf+xml",
                #                                        base_iri="http://iec.ch/TC57/2013/CIM-schema-cim16#")
                #
                # # pyoxigraph_iterator_list = list(pyoxigraph_iterator)
                # # data_frame = pl.DataFrame(pyoxigraph_iterator,
                # #                           schema=['triple'])  # this is possible but it is hard to split the triples
                #
                # data_for_graph = self.process_iterator(pyoxigraph_iterator)
                # # end Variant C
                #
                # end_time_parsing = time.time()  # end time parsing
                # elapsed_time_parsing = end_time_parsing - start_time_parsing
                #
                # print(f"Parsing time: {elapsed_time_parsing} seconds")
                # start_time_create_graph = time.time()  # start time create graph
                # # local_graph = Graph()
                # # local_graph.addN(data_for_graph)
                # self.merged_instance_graph = self.merged_instance_graph + data_for_graph
                # end_time_create_graph = time.time()  # end time create graph
                # elapsed_time_cr_graph = end_time_create_graph - start_time_create_graph
                # print(f"Creating graph time: {elapsed_time_cr_graph} seconds")

    def process_triple(self, triple, angle_bracket_pattern):
        s, p, o = triple
        match = angle_bracket_pattern.match(o)
        match_p = angle_bracket_pattern.match(p).group(1)

        if match:
            graph_object = rdflib.URIRef(match.group(1))
        else:
            filtered_df = self.datatypes_mapping.filter(self.datatypes_mapping['Property'] == match_p)
            if not filtered_df.is_empty():
                datatype_from_map = filtered_df[0]['Datatype']
            else:
                datatype_from_map = None

            if datatype_from_map is not None:
                graph_object = rdflib.Literal(o, datatype=rdflib.XSD.string)
            else:
                graph_object = rdflib.Literal(o, datatype=rdflib.URIRef(str(datatype_from_map[0, 1])))

        return rdflib.URIRef(angle_bracket_pattern.match(s).group(1)), rdflib.URIRef(match_p), graph_object

    def process_instance_data_contents(self, file_paths):
        for file_path in file_paths:
            with open(file_path, 'rb') as file:
                content = file.read()
                self.process_entry_content(file_path, content)

    def push_button_ok(self):  # button "OK"

        datatype_mapping = []
        instance_data_files = []
        shacl_file = []
        for action, details in self.file_types.items():
            if action == 'Instance Data':
                instance_data_files = details.get('selection', [])
            elif action == 'SHACL Constraints':
                shacl_file = details.get('selection', [])
            elif action == 'RDF Datatypes':
                datatype_mapping = details.get('selection', [])

        # import to Graph if the map is in .rdf
        # for file in datatype_mapping:
        #     local_datatype_mapping_graph = Graph()
        #     local_datatype_mapping_graph.parse(file, format='xml')
        #     self.merged_datatype_graph = self.merged_datatype_graph + local_datatype_mapping_graph

        # import from xlsx if the datatypes map is in xlsx
        for file in datatype_mapping:
            ModShape.datatypes_mapping = pl.read_excel(source=file, sheet_name="RDFS Datatypes")

        start_time_preparation = time.time()  # start time to prepare
        self.process_instance_data_contents(instance_data_files)
        end_time_preparation = time.time()  # end time to prepare
        elapsed_time_preparation = end_time_preparation - start_time_preparation
        print(f"Instance data parsing time: {elapsed_time_preparation} seconds")

        # load SHACL
        local_shacl_graph_data = Graph()
        for shacl_file in shacl_file:
            with open(shacl_file, 'rb') as file:
                content = file.read()
                sub_iterator = pyoxigraph.parse(input=content, mime_type="text/turtle")
                local_shacl_graph_data = self.process_shacl_iterator(sub_iterator, local_shacl_graph_data)
                self.merged_shacl_graph = self.merged_shacl_graph + local_shacl_graph_data

        # load SHACL variant 2
        # with ThreadPoolExecutor() as executor:
        #     futures = [executor.submit(self.load_owl_imports, file) for file in shacl_file]
        #
        #     for future in futures:
        #         local_shacl_data_graph = future.result()
        #         self.merged_shacl_graph += local_shacl_data_graph

        # #load SHACL variant 3
        # for file in shacl_file:
        #     local_shacl_data_graph = self.load_owl_imports(file, visited_files=None)
        #     self.merged_shacl_graph = self.merged_shacl_graph + local_shacl_data_graph

        # this below on the dataset is a trial. maybe useful when we do multiple graphs and combining graphs
        # instanceDataDataset = Dataset()  # 1st define the graph
        # instanceDataDataset.parse(instanceDataFile,format="xml",publicID=instanceDataDataset.default_context.identifier)
        #
        # shaclDataDataset = Dataset()  # 1st define the graph
        # shaclDataDataset.parse(shaclFile,format="turtle",publicID=shaclDataDataset.default_context.identifier)

        # as the inference somehow does not work, here another way to add the datatypes

        # start adding datatypes - version 1
        # graph_literals = self.merged_instance_graph.query(
        #     """SELECT DISTINCT ?p
        #         WHERE {
        #             ?s ?p ?o
        #             FILTER isLiteral(?o)
        #         }"""
        # )
        #
        # # then loop on the literals and if the literal is in tha datatype map then add teh datatype
        # for literal in graph_literals:
        #     datatype = self.merged_datatype_graph.triples((literal[0], RDFS.range, None))
        #     for s_d, p_d, o_d in datatype:
        #         # print(f"The datatype of {s_d} is {o_d}")
        #         for s_i, p_i, o_i in self.merged_instance_graph.triples((None, s_d, None)):
        #             # print(f"{s_i} has predicate {p_i} which is {o_i}")
        #             self.merged_instance_graph.add((s_i, p_i, rdflib.Literal(o_i,
        #                                                                      datatype=o_d.__str__())))  # this triggers a warning that can be ignorred
        #             self.merged_instance_graph.remove((s_i, p_i, o_i))
        # end adding datatypes - version 1

        start_time_validation = time.time()  # start time validation
        # this is validation without inference
        r = validate(self.merged_instance_graph,
                     shacl_graph=self.merged_shacl_graph,
                     ont_graph=None,
                     inference='none',
                     abort_on_first=False,
                     allow_infos=False,
                     allow_warnings=False,
                     meta_shacl=False,
                     advanced=False,
                     js=False,
                     debug=False,
                     do_owl_imports=False)

        conforms, results_graph, results_text = r

        end_time_validation = time.time()  # end time validation

        # Calculate the elapsed time
        elapsed_time_validation = end_time_validation - start_time_validation

        print(f"Validation time: {elapsed_time_validation} seconds")

        file_name = os.path.normpath(QFileDialog.getSaveFileName(self, "xxx", "C:", "*.xlsx")[0])

        # file_name = os.path.normpath(QFileDialog.getSaveFileName(self, "xxx", "C:", "*.jsonld")[0])
        # results_graph.serialize(destination=file_name, format='json-ld')
        wb = Workbook()
        ws = wb.active
        headers = ['Severity', 'Focus node', 'Path', 'Value', 'Message', 'Source shape']
        ws.append(headers)

        for s in results_graph.subjects(RDF.type, rdflib.term.URIRef("http://www.w3.org/ns/shacl#ValidationResult")):
            row = OrderedDict()
            for p in results_graph.predicates(s, None):
                if p == rdflib.term.URIRef("http://www.w3.org/ns/shacl#resultSeverity"):
                    row['Severity'] = results_graph.objects(s, p).__next__().__str__()
                elif p == rdflib.term.URIRef("http://www.w3.org/ns/shacl#focusNode"):
                    row['Focus node'] = results_graph.objects(s, p).__next__().__str__()
                elif p == rdflib.term.URIRef("http://www.w3.org/ns/shacl#resultPath"):
                    row['Path'] = results_graph.objects(s, p).__next__().__str__()
                elif p == rdflib.term.URIRef("http://www.w3.org/ns/shacl#value"):
                    row['Value'] = results_graph.objects(s, p).__next__().__str__()
                elif p == rdflib.term.URIRef("http://www.w3.org/ns/shacl#resultMessage"):
                    row['Message'] = results_graph.objects(s, p).__next__().__str__()
                elif p == rdflib.term.URIRef("http://www.w3.org/ns/shacl#sourceShape"):
                    row['Source shape'] = results_graph.objects(s, p).__next__().__str__()
            ws.append([row.get(header, '') for header in headers])

        # Format the worksheet
        for column_idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[self.get_column_letter(column_idx)].width = adjusted_width

        # Wrap text in a specific column (e.g., 'result_message')
        result_message_col = headers.index('Message') + 1
        for row in ws.iter_rows(min_col=result_message_col, max_col=result_message_col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # Add filter to the worksheet
        ws.auto_filter.ref = ws.dimensions

        # Freeze the first row
        ws.freeze_panes = 'A2'
        wb.save(file_name)

        print(f'{"Validation finished"}')
        print(f'Conforms?: {conforms}')

    @staticmethod
    def is_url(path):
        # Check if the path starts with "http://" or "https://"
        return path.startswith(("http://", "https://"))

    # Function to convert column index to letter
    def get_column_letter(self,col_idx):
        dividend = col_idx
        column_letter = ''
        while dividend > 0:
            modulo = (dividend - 1) % 26
            column_letter = chr(65 + modulo) + column_letter
            dividend = (dividend - modulo) // 26
        return column_letter

    def load_owl_imports(self, file_path, visited_files=None):
        # If visited_files is not provided, create a new set
        if visited_files is None:
            visited_files = set()
        # Create an RDF graph
        local_graph = Graph()

        # Parse the main file
        if self.is_url(file_path):
            # Fetch the file content from the URL
            response = requests.get(file_path)
            response.raise_for_status()
            local_graph.parse(data=response.text, format='turtle')
        else:
            # Parse the local file
            local_graph.parse(file_path, format='turtle')

        # Check for owl:imports triples
        imports = local_graph.objects(predicate=OWL.imports)

        for import_uri in imports:
            import_path = str(import_uri)

            # Avoid infinite recursion by checking if the file has already been visited
            if import_path not in visited_files:
                visited_files.add(import_path)

                # Parse the imported file
                imported_graph = self.load_owl_imports(import_path, visited_files)

                # Merge the imported graph into the main graph
                local_graph = local_graph + imported_graph

        return local_graph

        # if visited_files is None:
        #     visited_files = set()
        # local_graph = Graph()
        #
        # if self.is_url(file_path):
        #     response = requests.get(file_path)
        #     response.raise_for_status()
        #     local_graph.parse(data=response.text, format='turtle')
        # else:
        #     local_graph.parse(file_path, format='turtle')
        #
        # imports = local_graph.objects(predicate=OWL.imports)
        #
        # for import_uri in imports:
        #     import_path = str(import_uri)
        #
        #     if import_path not in visited_files:
        #         visited_files.add(import_path)
        #         imported_graph = self.load_owl_imports(import_path, visited_files)
        #         local_graph += imported_graph
        #
        # return local_graph


def main():
    qt_app = QApplication.instance()  # reuse old Qt application
    if qt_app is None:
        qt_app = QApplication(sys.argv)  # create a new Qt application

    exitcode = 1
    try:
        form = ModShape()
        form.show()
        exitcode = qt_app.exec_()
    finally:
        sys.exit(exitcode)


if __name__ == '__main__':
    main()
